"""Read the AppSec (SCA) worksheet exported from Tableau.

These are findings Nexus IQ has NOT raised as policy violations yet — quarantined or
soon-to-be-flagged libraries — so nothing in the IQ pipeline can discover them. There is no
API between Tableau and this tool, so the data arrives as a manually exported workbook and
this module is the only thing that reads it.

The export is dirty in ways that matter, and every case handled below was observed in a
real sheet rather than imagined:

  * `LIBRARY_NAME` is a human label ("Bouncy Castle Provider"), not a coordinate. All of the
    identity lives in `LIBRARY_FILENAME` — `bcprov-jdk15on-1.49.jar`.
  * `VULN_TOPFIX_RESOLUTION` is a LIST of candidates, separated by "," or ";" or ", ", and
    most of them name a DIFFERENT artifact than the one installed. See `parse_topfix`.
  * The same library repeats once per CVE. Eight of the first twelve rows of the sample were
    the same `bcprov-jdk15on-1.49.jar`. See `dedupe`.

Nothing here decides anything. Parsing produces facts and an explicit record of what it
could not read; choosing a target version is `appsec/resolve.py`.
"""

from __future__ import annotations

import logging
import re
from dataclasses import dataclass, field, replace
from datetime import date, datetime
from pathlib import Path

from nexus_autofix.iq.filter import is_a_real_upgrade

log = logging.getLogger(__name__)

#: Header -> internal field. Matched case-insensitively with whitespace collapsed, so
#: "CVSS3 Score", "cvss3  score" and "CVSS3_SCORE" all land in the same place. Overridable
#: from config.yml, because a Tableau re-export renames columns sooner or later.
DEFAULT_COLUMNS = {
    "github_org": "GITHUB_ORG",
    "github_repo_name": "GITHUB_REPO_NAME",
    "library_filename": "LIBRARY_FILENAME",
    "topfix": "VULN_TOPFIX_RESOLUTION",
    "library_name": "LIBRARY_NAME",
    "library_type": "LIBRARY_TYPE",
    "vuln_name": "VULN_NAME",
    "direct": "DIRECT_DEPENDENCY",
    "cvss3": "CVSS3 Score",
}

#: Without these four there is nothing to act on: no repo to attribute the row to, no
#: library to identify, and no suggested fix.
REQUIRED_FIELDS = ("github_org", "github_repo_name", "library_filename", "topfix")

#: Packaging suffixes stripped before pulling the version off a filename.
ARCHIVE_SUFFIXES = (".jar", ".war", ".ear", ".zip")

#: artifact-version, where the version must START with a digit. That requirement is what
#: makes the non-greedy artifact group land in the right place: for "bcprov-jdk15on-1.49"
#: the shortest match ("bcprov") leaves "jdk15on-1.49", which fails \d and backtracks to
#: the correct split. It is also what correctly REFUSES "jtidy-r938", where "r938" is a
#: revision rather than a version and guessing at one would be worse than reporting none.
_FILENAME_RE = re.compile(r"^(?P<artifact>.+?)-(?P<version>\d[0-9A-Za-z._\-]*)$")

_COORD = r"[A-Za-z0-9._\-]+"
#: The well-formed case: group:artifact:version.
_GAV_RE = re.compile(rf"^(?P<group>{_COORD}):(?P<artifact>{_COORD}):(?P<version>\d[0-9A-Za-z._\-]*)$")
#: Seen in the wild: "org.apache.servicemix.bundles:org.apache.servicemix.bundles.jdom - 2.0.5_1".
#: The version is separated by " - " rather than a colon. Supported because it appeared in a
#: 26-row sample, so it is not an outlier, and discarding it would silently lose a real
#: candidate.
_GAV_DASH_RE = re.compile(rf"^(?P<group>{_COORD}):(?P<artifact>{_COORD})\s+-\s+(?P<version>\d[0-9A-Za-z._\-]*)$")


@dataclass(frozen=True)
class Gav:
    group_id: str
    artifact_id: str
    version: str

    def __str__(self) -> str:
        return f"{self.group_id}:{self.artifact_id}:{self.version}"


@dataclass(frozen=True)
class AppsecRow:
    """One spreadsheet row: one CVE against one library in one repo."""

    github_org: str
    github_repo_name: str
    library_filename: str
    library_name: str
    library_type: str
    vuln_name: str
    artifact_id: str
    current_version: str
    direct: bool | None
    cvss3: float | None
    topfix: tuple[Gav, ...]
    #: Tokens in VULN_TOPFIX_RESOLUTION that could not be read as a coordinate. Kept rather
    #: than dropped so a candidate lost to a parsing gap is visible instead of silent.
    topfix_discarded: tuple[str, ...]


@dataclass(frozen=True)
class AppsecLibrary:
    """One library in one repo, with every row about it folded together."""

    github_org: str
    github_repo_name: str
    library_filename: str
    library_name: str
    library_type: str
    artifact_id: str
    current_version: str
    direct: bool | None
    cve_ids: tuple[str, ...]
    max_cvss3: float | None
    #: Highest candidate naming THIS artifact, and the group id it came with. None when the
    #: sheet only ever proposes other artifacts, or when it proposes this artifact under
    #: more than one group id — see `ambiguous_candidates`.
    sheet_version: str | None
    group_id: str | None
    #: Candidates naming a DIFFERENT artifact — a migration (bcprov-jdk15on -> bc-fips),
    #: not a version bump. Recorded, never applied. See resolve.py.
    swap_candidates: tuple[Gav, ...]
    topfix_discarded: tuple[str, ...]
    #: Same artifactId under DIFFERENT group ids, when the installed group is unknown.
    #: `com.fasterxml.jackson.core:jackson-core` and `tools.jackson.core:jackson-core` are
    #: not the same package, and LIBRARY_FILENAME does not say which one is installed.
    #: Picking the higher version would silently change the group — a migration disguised as
    #: a bump. Left for a human instead. Empty in the ordinary case.
    ambiguous_candidates: tuple[Gav, ...] = ()


@dataclass
class SheetStats:
    rows_total: int = 0
    rows_kept: int = 0
    skipped_missing_repo: int = 0
    skipped_missing_filename: int = 0
    skipped_unparsable_filename: list[str] = field(default_factory=list)
    skipped_wrong_type: int = 0


class SheetError(RuntimeError):
    """The workbook could not be read as an AppSec export."""


def _text(value: object) -> str:
    """Coerce a cell to text.

    Not cosmetic. `1.56` in VULN_TOPFIX_RESOLUTION arrives as a float, `COUNT(*)` as an
    int, and DUE_DATE as a datetime; calling .strip() on any of them raises.
    """
    if value is None:
        return ""
    if isinstance(value, (datetime, date)):
        return value.isoformat()
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _normalise_header(value: object) -> str:
    return re.sub(r"[\s_]+", " ", _text(value)).strip().lower()


def _as_bool(value: object) -> bool | None:
    """DIRECT_DEPENDENCY, which arrives as TRUE/FALSE text or a real bool."""
    if isinstance(value, bool):
        return value
    text = _text(value).lower()
    if text in ("true", "yes", "1"):
        return True
    if text in ("false", "no", "0"):
        return False
    return None


def _as_float(value: object) -> float | None:
    try:
        return float(_text(value))
    except ValueError:
        return None


def parse_filename(filename: str) -> tuple[str, str] | None:
    """Split a library filename into (artifact_id, version), or None if it has no version.

        bcprov-jdk15on-1.49.jar -> ("bcprov-jdk15on", "1.49")
        hsqldb-1.8.0.10.jar     -> ("hsqldb", "1.8.0.10")
        jtidy-r938.jar          -> None

    This is the ONLY source of component identity in the sheet. LIBRARY_NAME is a display
    label ("Bouncy Castle Provider") that matches neither a groupId nor an artifactId.
    """
    stem = _text(filename)
    for suffix in ARCHIVE_SUFFIXES:
        if stem.lower().endswith(suffix):
            stem = stem[: -len(suffix)]
            break
    match = _FILENAME_RE.match(stem)
    if not match:
        return None
    return match.group("artifact"), match.group("version")


def parse_topfix(cell: object) -> tuple[list[Gav], list[str]]:
    """Read VULN_TOPFIX_RESOLUTION into (usable coordinates, unreadable tokens).

    The column is a list of candidate remediations, and reading it as a single version is
    the mistake to avoid. A real cell, abbreviated:

        org.bouncycastle:bcprov-debug-jdk15on:1.56,org.bouncycastle:bc-fips:1.0.2;
        org.bouncycastle:bcprov-jdk15on:1.64, BouncyCastle.Cryptography

    Only the third of those names the artifact that is actually installed. Sorting the
    candidates by artifact is `dedupe`'s job; this function only decides what is a
    coordinate at all.

    Everything unreadable is RETURNED rather than dropped — bare versions ("1.56"), .NET
    package names, embedded URLs, and explicit non-answers ("jmeter - no_fix") all show up,
    and a caller that cannot see them cannot tell a clean row from a mangled one.
    """
    text = _text(cell)
    if not text:
        return [], []

    usable: list[Gav] = []
    discarded: list[str] = []
    for token in re.split(r"[,;]", text):
        token = token.strip()
        if not token:
            continue
        match = _GAV_RE.match(token) or _GAV_DASH_RE.match(token)
        if match:
            usable.append(Gav(match.group("group"), match.group("artifact"), match.group("version")))
        else:
            discarded.append(token)
    return usable, discarded


def _resolve_columns(header_row: tuple, overrides: dict[str, str] | None) -> dict[str, int]:
    """Map internal field -> column index, by header text."""
    wanted = {**DEFAULT_COLUMNS, **(overrides or {})}
    present = {_normalise_header(cell): index for index, cell in enumerate(header_row)}

    columns: dict[str, int] = {}
    for field_name, header in wanted.items():
        index = present.get(_normalise_header(header))
        if index is not None:
            columns[field_name] = index

    missing = [f for f in REQUIRED_FIELDS if f not in columns]
    if missing:
        # Names what was actually found. A re-export that renames a column is the failure
        # this will hit, and "GITHUB_ORG not found" without the alternatives leaves the
        # reader guessing which of thirty headers to map it to.
        raise SheetError(
            "the worksheet is missing required column(s): "
            + ", ".join(f"{f} (expected header {wanted[f]!r})" for f in missing)
            + "\n  Headers found: "
            + ", ".join(sorted(h for h in present if h))
            + "\n  Map them under `appsec.columns` in config.yml if the export renamed them."
        )
    return columns


def read_rows(
    path: Path,
    *,
    columns: dict[str, str] | None = None,
    library_type: str = "",
) -> tuple[list[AppsecRow], SheetStats]:
    """Read every usable row from the workbook's first sheet.

    `library_type` filters on LIBRARY_TYPE ("Java"). Rows for another ecosystem are counted
    and skipped rather than guessed at — an npm filename does not split the way a jar does.
    """
    try:
        import openpyxl
    except ImportError as exc:  # pragma: no cover - dependency is declared
        raise SheetError("openpyxl is required to read the AppSec worksheet") from exc

    # data_only so formula cells yield their computed value rather than "=...".
    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        sheet = workbook[workbook.sheetnames[0]]
        rows = sheet.iter_rows(values_only=True)
        try:
            header_row = next(rows)
        except StopIteration:
            raise SheetError(f"{path} has no rows at all") from None

        index = _resolve_columns(header_row, columns)
        stats = SheetStats()
        parsed: list[AppsecRow] = []

        def cell(row: tuple, field_name: str) -> object:
            position = index.get(field_name)
            if position is None or position >= len(row):
                return None
            return row[position]

        for row in rows:
            if row is None or not any(c is not None for c in row):
                continue
            stats.rows_total += 1

            org = _text(cell(row, "github_org"))
            repo = _text(cell(row, "github_repo_name"))
            if not org or not repo:
                stats.skipped_missing_repo += 1
                continue

            filename = _text(cell(row, "library_filename"))
            if not filename:
                stats.skipped_missing_filename += 1
                continue

            row_type = _text(cell(row, "library_type"))
            if library_type and row_type and row_type.lower() != library_type.lower():
                stats.skipped_wrong_type += 1
                continue

            split = parse_filename(filename)
            if split is None:
                # No version in the filename means no way to tell an upgrade from a
                # downgrade later, so this cannot be acted on. Named, not silently dropped.
                stats.skipped_unparsable_filename.append(filename)
                continue
            artifact_id, current_version = split

            usable, discarded = parse_topfix(cell(row, "topfix"))
            parsed.append(
                AppsecRow(
                    github_org=org,
                    github_repo_name=repo,
                    library_filename=filename,
                    library_name=_text(cell(row, "library_name")),
                    library_type=row_type,
                    vuln_name=_text(cell(row, "vuln_name")),
                    artifact_id=artifact_id,
                    current_version=current_version,
                    direct=_as_bool(cell(row, "direct")),
                    cvss3=_as_float(cell(row, "cvss3")),
                    topfix=tuple(usable),
                    topfix_discarded=tuple(discarded),
                )
            )
            stats.rows_kept += 1
    finally:
        workbook.close()

    log.info(
        "AppSec sheet %s: %d row(s), %d kept, %d skipped (no repo=%d, no filename=%d, "
        "wrong type=%d, unreadable filename=%d)",
        path, stats.rows_total, stats.rows_kept,
        stats.rows_total - stats.rows_kept, stats.skipped_missing_repo,
        stats.skipped_missing_filename, stats.skipped_wrong_type,
        len(stats.skipped_unparsable_filename),
    )
    return parsed, stats


def _highest(versions: list[str]) -> str | None:
    """The greatest of a list of versions, by the same ordering used everywhere else."""
    best: str | None = None
    for version in versions:
        if best is None or is_a_real_upgrade(best, version):
            best = version
    return best


def dedupe(rows: list[AppsecRow]) -> list[AppsecLibrary]:
    """Fold the per-CVE rows into one entry per (repo, library).

    The export repeats a library once per CVE — eight of the first twelve rows of the
    sample were the same `bcprov-jdk15on-1.49.jar`, each with a different suggested fix.
    Treating those as eight findings would hand the agent eight conflicting instructions
    for one line of one manifest.

    Candidates are split by artifact. Same artifact is a version bump this tool can make
    and verify; a different artifact is a migration (bcprov-jdk15on -> bc-fips) that the
    diff classifier would refuse anyway, so it is recorded for a human and never applied.
    The highest same-artifact version wins — of several versions each clearing one CVE, the
    highest is the one most likely to clear all of them.
    """
    grouped: dict[tuple[str, str, str], list[AppsecRow]] = {}
    for row in rows:
        key = (row.github_org.lower(), row.github_repo_name.lower(), row.library_filename.lower())
        grouped.setdefault(key, []).append(row)

    libraries: list[AppsecLibrary] = []
    for group in grouped.values():
        first = group[0]
        artifact = first.artifact_id.lower()

        same_artifact = [g for row in group for g in row.topfix if g.artifact_id.lower() == artifact]
        swaps = [g for row in group for g in row.topfix if g.artifact_id.lower() != artifact]

        # A matching artifactId is NOT enough. "com.fasterxml.jackson.core:jackson-core"
        # and "tools.jackson.core:jackson-core" share an artifact name and are different
        # packages; taking the higher version across both would swap the group id silently,
        # which is the migration this code exists to refuse. LIBRARY_FILENAME never states
        # a group, so when the candidates disagree about it there is nothing here that can
        # tell which is installed — `narrow_to_group` resolves it later if Nexus IQ knows,
        # and otherwise it goes to a human.
        groups = {g.group_id.lower() for g in same_artifact}
        ambiguous = tuple(dict.fromkeys(same_artifact)) if len(groups) > 1 else ()

        usable = [] if ambiguous else same_artifact
        sheet_version = _highest([g.version for g in usable])
        # The group id belongs to the candidate that won, and is the only place the sheet
        # states one: LIBRARY_FILENAME carries the artifact and version but never the group.
        group_id = next((g.group_id for g in usable if g.version == sheet_version), None)

        cvss = [row.cvss3 for row in group if row.cvss3 is not None]
        # dict.fromkeys rather than set(): order follows the sheet, so the CVE list in
        # run.json is stable across runs and diffs cleanly.
        cve_ids = tuple(dict.fromkeys(row.vuln_name for row in group if row.vuln_name))
        direct = next((row.direct for row in group if row.direct is not None), None)

        libraries.append(
            AppsecLibrary(
                github_org=first.github_org,
                github_repo_name=first.github_repo_name,
                library_filename=first.library_filename,
                library_name=first.library_name,
                library_type=first.library_type,
                artifact_id=first.artifact_id,
                current_version=first.current_version,
                direct=direct,
                cve_ids=cve_ids,
                max_cvss3=max(cvss) if cvss else None,
                sheet_version=sheet_version,
                group_id=group_id,
                swap_candidates=tuple(dict.fromkeys(swaps)),
                topfix_discarded=tuple(dict.fromkeys(t for row in group for t in row.topfix_discarded)),
                ambiguous_candidates=ambiguous,
            )
        )
    return libraries


def narrow_to_group(library: AppsecLibrary, group_id: str) -> AppsecLibrary:
    """Resolve an ambiguous library once the real group id is known.

    Nexus IQ's policy report states the installed groupId, which the sheet never does. When
    a library was left ambiguous because its candidates spanned several groups, that answer
    settles it: candidates under the real group are version bumps, and the rest are
    migrations to a different package and belong with the swaps.

    A no-op for a library that was never ambiguous.
    """
    if not library.ambiguous_candidates:
        return library

    wanted = group_id.lower()
    matching = [g for g in library.ambiguous_candidates if g.group_id.lower() == wanted]
    other = [g for g in library.ambiguous_candidates if g.group_id.lower() != wanted]
    if not matching:
        # Every candidate is under a different group than what is installed: all migrations.
        return replace(
            library, ambiguous_candidates=(),
            swap_candidates=tuple(dict.fromkeys([*library.swap_candidates, *other])),
        )

    version = _highest([g.version for g in matching])
    return replace(
        library,
        sheet_version=version,
        group_id=next(g.group_id for g in matching if g.version == version),
        ambiguous_candidates=(),
        swap_candidates=tuple(dict.fromkeys([*library.swap_candidates, *other])),
    )


def for_repo(libraries: list[AppsecLibrary], owner: str, repo: str) -> list[AppsecLibrary]:
    """Just this repository's libraries.

    GITHUB_ORG / GITHUB_REPO_NAME are exactly what `_owner_repo_from_url` pulls out of a
    clone URL in config.yml, so the join needs no configuration of its own.
    """
    return [
        library
        for library in libraries
        if library.github_org.lower() == owner.lower()
        and library.github_repo_name.lower() == repo.lower()
    ]
